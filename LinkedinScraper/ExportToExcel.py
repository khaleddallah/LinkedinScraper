#! /usr/bin/env python

# Author : Khaled Dallah
# Email : khaled.dallah0@gmail.com
# Date : 8-12-2018


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

from shutil import copyfile
from copy import copy
import json, os



class Ete :
	#current Available row to load data in it
	rowIndex=3
	picture=''
	header={
	'fs_profile':[{'firstName':'', 'lastName':'', 'locationName':'', 'headline':'','summary':''}],
	'fs_position':[{'title':'', 'companyName':'', 'locationName':'' }],
	'fs_education':[{'degreeName':'', 'schoolName':'', 'fieldOfStudy':'', 'activities':''}],
	'fs_volunteerExperience':[{'companyName':'', 'role':'', 'cause':''}],
	'fs_skill':[{'name':''}],
	'fs_certification':[{'authority':'', 'name':'', 'licenseNumber':''}],
	'fs_course':[{'name':''}],
	'fs_language':[{'name':'','proficiency':''}],
	'fs_project':[{'title':'', 'description':'', 'url':''}],
	'fs_honor':[{'description':'', 'title':'', 'issuer':''}]
	}

	#Set the Input & Output 	
	def __init__(self,name,data,mode):
		directory='Output'
		if not os.path.exists(directory):
			os.makedirs(directory)
		self.filePath=directory+'/'+name+'.xlsx'
		self.data=data
		self.mode=mode


	#Open Excel file and first sheet 
	def open_file_sheet(self):
		print('excel file path : ',self.filePath)
		copyfile('Template/atemp.xlsx',self.filePath)
		self.wb1 = load_workbook(self.filePath)
		self.ws1 = self.wb1['first']
		self.dh=[s for s in self.ws1[2]]

	#Save Excel file
	def save_file(self):
		self.wb1.save(filename = self.filePath)


	#Not useful after we use template
	#Write header data to Exel file
	def write_header(self):
		#additon for merge cell colomn
		amc=1
		for i in self.header:
			temp=self.ws1.cell(column=amc , row=1 , value=i[3:] )
			temp.alignment=self.al
			shift=len(self.header[i][0])
			self.ws1.merge_cells(start_row=1, start_column=amc, end_row=1, end_column=amc+shift-1)
			#print("\n... Printing ",i,"in col ",amc,' and make marge from ',amc,' to ',amc+shift-1)
			subind=0
			for j in self.header[i][0]:
				temp=self.ws1.cell(column=amc+subind , row=2 , value=j )
				temp.alignment=self.al
				subind+=1
			amc+=shift


	

	#Function Data Loader
	def single_data_loader(self,sdata):
		fill= PatternFill("solid", fgColor="DDDDDD")
		border1=copy(self.ws1['A2'].border)
		shiftRow=0
		maxShiftRow=0
		for i in sdata:
			#to process picture alone
			if (i=='fs_miniProfile'):
				self.picture=sdata[i]
				continue

			shiftRow=0
			currentSec=i
			for j in sdata[i]:
				shiftRow+=1
				for k in j:
					#Get the right colomn
					tempCol=self.getRightCol(k,currentSec)
					temp=self.ws1.cell(column=tempCol , row=self.rowIndex+shiftRow , value=j[k])
					self.copyStyle(temp,self.dh[tempCol-1])

			if(shiftRow>=maxShiftRow):
				maxShiftRow=shiftRow

		#Shadow all rows that profile use
		for rc in range(self.rowIndex+1, self.rowIndex+maxShiftRow+1):
			for cc in range(len(self.ws1[rc])):
				curcell=self.ws1.cell(column=cc+1, row=rc)
				if(curcell.value):
					continue
				else:
					curcell.fill=fill
					curcell.border = border1

		self.rowIndex+=1+maxShiftRow


	#Function Data Loader for one row
	def single_data_loader_in_one_row(self,sdata):
		for i in sdata:
			#to process picture alone
			if (i=='fs_miniProfile'):
				self.picture=sdata[i]
				continue
			currentSec=i
			for j in sdata[i]:
				for k in j:
					#Get the right colomn
					tempCol=self.getRightCol(k,currentSec)
					temp1=self.ws1.cell(column=tempCol , row=self.rowIndex)
					if (str(temp1.value)=='None'):
						temp2=''
					else:	
						temp2=str(temp1.value)+','
					temp=self.ws1.cell(column=tempCol , row=self.rowIndex, value=temp2+j[k])
					self.copyStyle(temp,self.dh[tempCol-1])
		self.rowIndex+=1



	#to parse all Data
	def all_data_loader(self):
		for i in self.data:
			if(self.mode=='1'):
				self.single_data_loader_in_one_row(i)
			elif(self.mode=='m'):
				self.single_data_loader(i)
			else:
				self.single_data_loader(i)



	def copyStyle(self,target,src):
		target.border = copy(src.border)
		target.fill = copy(src.fill)
		target.number_format = copy(src.number_format)
		target.protection = copy(src.protection)
		target.alignment = copy(src.alignment)		



	#to put item value data in right cell
	def getRightCol(self,item,currentSec):
		t1=False
		t2=False
		ind=0
		for i in self.header:
			if (i==currentSec):
				t1=True
			for j in self.header[i][0]:
				ind+=1
				if(j==item):
					t2=True
					if(t1):
						return(ind)

		else:
			print('\n!!! ERROR to find <',item,'>\n')
			return(-1)


	#coloring data
	def coloring(self):
		pass


	#Run Basic Operation
	def run(self):
		self.open_file_sheet()
		self.all_data_loader()
		self.save_file()




def main():
	pathh=''
	with open(pathh,'r+') as f:
		jsonData=json.load(f)


	ete1=Ete('aa.xlsx',jsonData)
	ete1.run()


if __name__=='__main__':
	main()